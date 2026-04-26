"""Query-bank loader for typed planning tests and offline evaluation."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import Field

from strap.planning.capability_registry import ARTIFACT_TYPES
from strap.planning.models import PlanningModel


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_QUERY_BANK_PATH = PROJECT_ROOT / "docs" / "subagent_query_bank-v1.xlsx"


class QueryBankRow(PlanningModel):
    sheet_name: str
    role: str
    row_number: int
    query: str
    status: str | None = None
    priority: str | None = None
    focus: str | None = None
    expected_route_or_subagents: str | None = None
    expected_tools_or_handoffs: str | None = None
    required_inputs: str | None = None
    expected_outputs_or_artifacts: str | None = None
    validation_checks: str | None = None
    last_run_or_artifact_path: str | None = None
    notes: str | None = None
    raw: dict[str, Any] = Field(default_factory=dict)

    @property
    def is_validated(self) -> bool:
        return (self.status or "").strip().lower() == "validated"

    @property
    def is_p0(self) -> bool:
        return (self.priority or "").strip().upper() == "P0"

    def expected_tool_names(self, exported_tool_names: set[str] | frozenset[str]) -> list[str]:
        haystack = " ".join(
            item or ""
            for item in (
                self.expected_tools_or_handoffs,
                self.validation_checks,
                self.notes,
            )
        )
        found = [
            tool_name
            for tool_name in sorted(exported_tool_names, key=len, reverse=True)
            if re.search(rf"(?<![A-Za-z0-9_]){re.escape(tool_name)}(?![A-Za-z0-9_])", haystack)
        ]
        return sorted(set(found))

    def expected_artifact_types(self, artifact_types: set[str] | frozenset[str] = ARTIFACT_TYPES) -> list[str]:
        haystack = " ".join(
            item or ""
            for item in (
                self.expected_outputs_or_artifacts,
                self.validation_checks,
                self.notes,
            )
        ).lower()
        aliases = {
            "single solvent safety card": "solvent_safety_card",
            "single solvent card": "solvent_safety_card",
            "safety card": "solvent_safety_card",
            "comparison table": "solvent_safety_comparison",
            "hsp_binary_screen": "hsp_single_pair_summary",
            "single-pair radar/red": "hsp_single_pair_summary",
            "radar/red summary": "hsp_single_pair_summary",
            "structured error with error_code=unsupported_hsp_solvent": "hsp_single_pair_summary",
            "structured error with error_code=ambiguous_hsp_polymer": "hsp_single_pair_summary",
            "red heatmap": "hsp_red_heatmap",
            "structured point optimum": "optimization_point_result",
            "structured optimum": "optimization_point_result",
            "selected solvents": "optimization_point_result",
            "optimization figure": "optimization_point_plot",
            "pareto payload": "optimization_pareto_front",
            "all feasible points": "optimization_pareto_landscape",
            "pareto landscape": "optimization_pareto_landscape",
            "highlighted frontier": "optimization_pareto_front",
            "rich pareto landscape": "optimization_pareto_landscape",
            "one png per composition": "optimization_pareto_slices_plot",
            "combined comparison plot": "optimization_pareto_slices_plot",
            "sidecar json": "sidecar_file",
            "separation plots": "separation_topk_sequences",
        }
        found = [
            artifact_type
            for artifact_type in sorted(artifact_types, key=len, reverse=True)
            if re.search(rf"(?<![A-Za-z0-9_]){re.escape(artifact_type.lower())}(?![A-Za-z0-9_])", haystack)
        ]
        for phrase, artifact_type in aliases.items():
            if phrase in haystack and artifact_type in artifact_types:
                found.append(artifact_type)
        return sorted(set(found))


def normalize_role_from_sheet(sheet_name: str) -> str:
    role = re.sub(r"^\d+\s+", "", sheet_name.strip())
    if role == "contaminant-removal":
        return "contaminant-removal-analyst"
    return role


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_query_bank(path: str | Path = DEFAULT_QUERY_BANK_PATH) -> list[QueryBankRow]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[QueryBankRow] = []
    for worksheet in workbook.worksheets:
        values_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = [_normalize_header(value) for value in next(values_iter)]
        except StopIteration:
            continue
        for row_number, values in enumerate(values_iter, start=2):
            raw = dict(zip(headers, values, strict=False))
            query = _string_or_none(raw.get("query"))
            if not query:
                continue
            rows.append(
                QueryBankRow(
                    sheet_name=worksheet.title,
                    role=normalize_role_from_sheet(worksheet.title),
                    row_number=row_number,
                    query=query,
                    status=_string_or_none(raw.get("status")),
                    priority=_string_or_none(raw.get("priority")),
                    focus=_string_or_none(raw.get("focus")),
                    expected_route_or_subagents=_string_or_none(raw.get("expected_route_or_subagents")),
                    expected_tools_or_handoffs=_string_or_none(raw.get("expected_tools_or_handoffs")),
                    required_inputs=_string_or_none(raw.get("required_inputs")),
                    expected_outputs_or_artifacts=_string_or_none(raw.get("expected_outputs_or_artifacts")),
                    validation_checks=_string_or_none(raw.get("validation_checks")),
                    last_run_or_artifact_path=_string_or_none(raw.get("last_run_or_artifact_path")),
                    notes=_string_or_none(raw.get("notes")),
                    raw={key: _string_or_none(value) for key, value in raw.items()},
                )
            )
    return rows


def validated_query_bank_rows(
    path: str | Path = DEFAULT_QUERY_BANK_PATH,
    *,
    priority: str | None = None,
) -> list[QueryBankRow]:
    rows = [row for row in load_query_bank(path) if row.is_validated]
    if priority is None:
        return rows
    expected = priority.strip().upper()
    return [row for row in rows if (row.priority or "").strip().upper() == expected]
